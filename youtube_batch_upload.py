import os

from tkinter import Tk, ttk
from tkinter import OptionMenu, Label, Button, GROOVE, Entry
from tkinter import StringVar
from tkinter import Toplevel
import sys
from tkinter import filedialog
from tkinter.ttk import *
from tkinter import *


abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)

'''
The following function allows to extract all filenames in provided path. The path should be provided in following from: "c:\folder\"
'''
folder_was_selected = 0

def filename_extractor(directory_name):
    file_names = []
    for root, dirs, files in os.walk(directory_name):
        for filename in files:
            file_names.append((filename, root))
    return file_names
    

#The following function allows to separate name lists of video, subtitle and image files out of all file names

def identify_extension(result_of_filename_extractor, extensions):
    results = []
    for any_file, root in result_of_filename_extractor:
        for extension in extensions:
            if extension in any_file.lower():
                results.append((any_file, root))
                break
    return results

def identify_videos(result_of_filename_extractor):
    extensions = [
        ".mp4",
        ".avi"
    ]
    return identify_extension(result_of_filename_extractor, extensions)

def identify_subtitles(result_of_filename_extractor):
    extensions = [
        ".srt",
        ".txt"
    ]
    return identify_extension(result_of_filename_extractor, extensions)

def identify_images(result_of_filename_extractor):
    extensions = [
        ".jpg",
        ".png"
    ]
    return identify_extension(result_of_filename_extractor, extensions)

#The following block of code defines functions that will adjust the filenames to our template:

def filename_matching(filenames, inclusion_patterns):
    result = ""
    for any_file, root in filenames:
        for pattern in inclusion_patterns: 
            if pattern in any_file.lower().replace("-","").replace("_",""):
                #if filename_exceptions(any_file.lower().replace("-","").replace("_","")) == False:
                result=any_file
    return result

def filename_exceptions(exception):
       return "dede" in exception and "pegi" in exception


# Image size check

def check_image_size(images_file_names):
    for filename, root in images_file_names:
        os.stat(root +'/'+ filename).st_size
        if os.stat(root +'/'+ filename).st_size > 2000000:
            image_too_big_message = 'This image is too big:'
            image_too_big_filename = filename
            image_too_big_explanation = 'YouTube thumbnails needs to be smaller than 2 MB.'
            insert_log(image_too_big_message)
            insert_log(image_too_big_filename)
            insert_log("")
            insert_log(image_too_big_explanation)

# User Input

selected_path = ""
all_files = {}
all_videos = {}
all_subtitles = {}
all_images = {}
video = {}
subtitle = {}
image = {}

'''Short videos for faster upload: \\eu.blizzard.net\Teams\Marketing\Web\Team\Editorial\Diablo III\2018-09-18_D3_Switch-PaidMedia\videos\PreLaunch\bumper'''



# Segregating videos, subtitles and images:

file_language = {
    "de-DE": {
        "caption_language": "de-DE",
        "subtitle_and_image_pattern": ["dede", "dedeusk"],
        "video_pattern": ["dede", "dedeusk"],
    },
    "de-DE-PEGI": {
        "caption_language": "de-DE",
        "subtitle_and_image_pattern": ["dede", "dedepegi"],
        "video_pattern": ["dedepegi"],
    },
    "en-US": {
        "caption_language": "en-US",
        "subtitle_and_image_pattern": ["engb","enus"],
        "video_pattern": ["enus"],
    },
    "en-GB": {
        "caption_language": "en-GB",
        "subtitle_and_image_pattern": ["enus","engb","eneu"],
        "video_pattern": ["engb", "eneu", "enen"],
    },
    "en-EU": {
        "caption_language": ["en-GB","en-EU"],
        "subtitle_and_image_pattern": ["eneu"],
        "video_pattern": ["eneu"],
    },
    "en-ANZ": {
        "caption_language": "en-US",
        "subtitle_and_image_pattern": ["engb","enus","enanz","enau"],
        "video_pattern": ["enanz","enau"],
    },
    "es-ES": {
        "caption_language": "es-ES",
        "subtitle_and_image_pattern": ["eses"],
        "video_pattern": ["eses"],
    },
    "es-MX": {
        "caption_language": "es-MX",
        "subtitle_and_image_pattern": ["esmx"],
        "video_pattern": ["esmx"],
    },
    "fr-FR": {
        "caption_language": "fr-FR",
        "subtitle_and_image_pattern": ["frfr"],
        "video_pattern": ["frfr"],
    },
    "it-IT": {
        "caption_language": "it",
        "subtitle_and_image_pattern": ["itit"],
        "video_pattern": ["itit"],
    },
    "ja-JP": {
        "caption_language": "ja",
        "subtitle_and_image_pattern": ["jajp","jpjp"],
        "video_pattern": ["jajp","jpjp"],
    },
    "ko-KR": {
        "caption_language": "ko",
        "subtitle_and_image_pattern": ["kokr"],
        "video_pattern": ["kokr"],
    },
    "pl-PL": {
        "caption_language": "pl",
        "subtitle_and_image_pattern": ["plpl"],
        "video_pattern": ["plpl"],
    },
    "pt-BR": {
        "caption_language": "pt-BR",
        "subtitle_and_image_pattern": ["ptbr"],
        "video_pattern": ["ptbr"],
    },
    "pt-PT": {
        "caption_language": "pt-PT",
        "subtitle_and_image_pattern": ["ptpt"],
        "video_pattern": ["ptpt"],
    },
    "ru-RU": {
        "caption_language": "ru",
        "subtitle_and_image_pattern": ["ruru"],
        "video_pattern": ["ruru"],
    },
    "th-TH": {
        "caption_language": "th",
        "subtitle_and_image_pattern": ["thth"],
        "video_pattern": ["thth"],
    },
    "zh-TW": {
        "caption_language": "zh-TW",
        "subtitle_and_image_pattern": ["zhtw"],
        "video_pattern": ["zhtw"],
    },
}

def video_dictionary():
    for languages in file_language:
        language_pattern = file_language[languages]["video_pattern"]
        video[languages] = filename_matching(all_videos, language_pattern)


def subtitle_and_image_dictionary():
    for languages in file_language:
        language_pattern = file_language[languages]["subtitle_and_image_pattern"]
        subtitle[languages] = filename_matching(all_subtitles, language_pattern)
        image[languages] = filename_matching(all_images, language_pattern)

# Counter for the number of videos

def filter_video(raw_list):
    list_of_elements = []
    for element in raw_list:
        if element == '':
            continue
        list_of_elements.append(element)
    return list_of_elements  

# Subtitle and image check:

def check_if_subtitles_exist():
    return any(subtitle.values())

def check_if_thumbnails_exists():
    return any(image.values())

# validation check for subs and images

def check_if_videos_subtitles_and_images_match():
    list_of_missing_subtitles = []
    list_of_missing_images = []
    for languages in file_language:
        if video[languages] and not subtitle[languages]:
            list_of_missing_subtitles.append(video[languages])
        if video[languages] and not image[languages] and (check_if_thumbnails_exists() or len(all_images) == 0):
            list_of_missing_images.append(video[languages])
        if len(all_videos) == 0:
            insert_log('There are no videos in the folder.')
            break
    if len(list_of_missing_subtitles) > 0:
        insert_log('Videos missing subtitles: ')
        insert_log("")
        for video_name in list_of_missing_subtitles:
            insert_log(video_name)
        insert_log("")
    if len(list_of_missing_images) > 0:
        insert_log('Videos missing images: ')
        insert_log("")
        for video_name in list_of_missing_images:
            insert_log(video_name)
        insert_log("")

# YT channel links:

franchises = {
    "Hearthstone": {
        "de-DE-chan": {
            "id": "UCs-3oulQFuVYV2nBAU0abyA",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCs-3oulQFuVYV2nBAU0abyA",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UCVia_crjzJylRmGq7SHTiaw",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UCVia_crjzJylRmGq7SHTiaw",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCVia_crjzJylRmGq7SHTiaw",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UCVia_crjzJylRmGq7SHTiaw",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UCfT5gqH_v9jUsJ4sR1B3mUA",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCEI3Et0HlPasooNL16qncRA",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UC1rWX8SuXh0DzHVKTfl5M0g",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UCtNW3-ZMKsTqmi4xl7HqjyQ",
            "languages": ["it-IT"]
        },
        "ja-JP-chan":{
            "id": "UCrR37g6dpxX4kUitZR8b4YA",
            "languages": ["ja-JP"]
        },
        "ko-KR-chan":{
            "id": "UCHyMBtXBR3uJyPg42jTwUgA",
            "languages": ["ko-KR"]
        },
        "pl-PL-chan":{
            "id": "UCn808LZdOGcGD0eC1_uJpyg",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UCiXMMud2GemSR8PhXhTHuyQ",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UCYBmtLck4mwBALwXidoC5Cw",
            "languages": ["ru-RU"]
        },
        "th-TH-chan":{
            "id": "UCID9zD5KaH6yG1pGFEE1Mog",
            "languages": ["th-TH"]
        },
        "zh-TW-chan":{
            "id": "UC4y1Cab73KDMb82ZDfu1QnQ",
            "languages": ["zh-TW"]
        },
    },
    "Heroes": {
        "de-DE-chan": {
            "id": "UC5j-bG5YieOsSE7sGlvVOkQ",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UC5j-bG5YieOsSE7sGlvVOkQ",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UCpVdq9gLew6E76BmfB2GJ0w",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UCpVdq9gLew6E76BmfB2GJ0w",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCpVdq9gLew6E76BmfB2GJ0w",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UCpVdq9gLew6E76BmfB2GJ0w",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UC1oU6m1xajKd05zQ-Meroqg",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCsJ5di4_Ie8QBfIirAW13LQ",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UCoFL3l9LQd9kutPt-OiJB8A",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UC2aFymYNAADfOqupz8Klk6g",
            "languages": ["it-IT"]
        },
        "ko-KR-chan":{
            "id": "UCSxwRv0puT0oW0Dic1SNW0w",
            "languages": ["ko-KR"]
        },
        "pl-PL-chan":{
            "id": "UC75IsW5SqCK28hOJeeti2Ww",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UCcdH9atAxgdbtPaUaGafTlg",
            "languages": ["pt-BR"]
        },
        "pt-PT-chan":{
            "id": "UCcdH9atAxgdbtPaUaGafTlg",
            "languages": ["pt-PT"]
        },
        "ru-RU-chan":{
            "id": "UCAD-mtX5cVUlvRJ6Z1P56Vg",
            "languages": ["ru-RU"]
        },
        "zh-TW-chan":{
            "id": "UC2t6HGb5YX1N8B9mp3HBcBA",
            "languages": ["zh-TW"]
        },
    },
    "Overwatch": {
        "de-DE-chan": {
            "id": "UCPY4lCVrQ13SXGlCgFQGXDA",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCPY4lCVrQ13SXGlCgFQGXDA",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UClOf1XXinvZsy4wKPAkro2A",
            "languages": ["en-US"]
        },
        "en-ANZ-chan": {
            "id": "UClOf1XXinvZsy4wKPAkro2A",
            "languages": ["en-ANZ"]
        },
        "en-GB-chan": {
            "id": "UCIUG4IllEehwwJMdeM9ejnQ",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCIUG4IllEehwwJMdeM9ejnQ",
            "languages": ["en-EU"]
        },
        "es-ES-chan": {
            "id": "UCbnHloj_Olxb1aFwRcCPklQ",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCbnHloj_Olxb1aFwRcCPklQ",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UC8fDAVEut58sjBXMiU76yQw",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UCtZSgPKKL4JG2Sdm4NQLOtQ",
            "languages": ["it-IT"]
        },
        "ja-JP-chan":{
            "id": "UCUlU_aydFO4u0cDDng1aO7w",
            "languages": ["ja-JP"]
        },
        "ko-KR-chan":{
            "id": "UC-2wa6jvprl7hfCpvw0ULzg",
            "languages": ["ko-KR"]
        },
        "pl-PL-chan":{
            "id": "UCqjHsVw9tI_KYdb4_duF7Vw",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UCcOlGZMY19XQupKH4aigvkw",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UCpW84gDcZu8wNQ-tUO5qE6A",
            "languages": ["ru-RU"]
        },
        "zh-TW-chan":{
            "id": "UCwAbpk7Jq5ursZLJXCDULnQ",
            "languages": ["zh-TW"]
        },
    },
    "Diablo": {
        "de-DE-chan": {
            "id": "UCOFtRcsBzVO3zUMVfu65yWw",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCOFtRcsBzVO3zUMVfu65yWw",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UCxn8csYeZg6awRnZS-aqg0g",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UCxn8csYeZg6awRnZS-aqg0g",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCxn8csYeZg6awRnZS-aqg0g",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UCxn8csYeZg6awRnZS-aqg0g",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UCuwx1Io52q8euJvpBLuPVjA",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCsV0O69egX50HKgpB7sphjA",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UCZazROYfSdlhf7zdFbL-xCQ",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UCd5eSodHP0mmx-zOigm3E5w",
            "languages": ["it-IT"]
        },
        "pl-PL-chan":{
            "id": "UCstAHhPqo_OXkl1YSAVIFBQ",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UC0vD7GP6_mkRJqkUMwo6L9w",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UCak7pETEvlckk9xCto28IEA",
            "languages": ["ru-RU"]
        },
        "zh-TW-chan":{
            "id": "UCRfXA8mX101VaAxT3CL-m-g",
            "languages": ["zh-TW"]
        },
    },
    "Starcraft": {
        "de-DE-chan": {
            "id": "UCj8yjTfEPZ4aafqI7uU4VPg",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCj8yjTfEPZ4aafqI7uU4VPg",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UCZ2h0d4tP0R8klduwj-_bDQ",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UCZ2h0d4tP0R8klduwj-_bDQ",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCZ2h0d4tP0R8klduwj-_bDQ",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UCZ2h0d4tP0R8klduwj-_bDQ",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UC7b-EG8HU8X8E7gQDaCfx5Q",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCn7fau6gDGzGpzFakk3U8eA",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UCbvO71zGR3uv6k3-PLSYzdQ",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UCs3-N9BsZwbh9U0VZduy4xA",
            "languages": ["it-IT"]
        },
        "pl-PL-chan":{
            "id": "UCWO3UQumwFAZCwd3DvtCx2A",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UCCNtLNS8N6AF9Zyjv0nAFig",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UCNXaW6faJ7mg-WkdMiD843Q",
            "languages": ["ru-RU"]
        },
        "zh-TW-chan":{
            "id": "UCLQRUOUYGwO6-YT_ROX2oNA",
            "languages": ["zh-TW"]
        },
    },
    "WoW": {
        "de-DE-chan": {
            "id": "UCOMpcd47VxMnodSCxYnw_yQ",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCOMpcd47VxMnodSCxYnw_yQ",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UCbLj9QP9FAaHs_647QckGtg",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UCbLj9QP9FAaHs_647QckGtg",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UCbLj9QP9FAaHs_647QckGtg",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UCbLj9QP9FAaHs_647QckGtg",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UCi5ZbDLYAWw0VXifoiVatAQ",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCuVY7xEIjQN6yE3emKRzq8g",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UCIYdx-MXGD8CjyLfDCfS-cQ",
            "languages": ["fr-FR"]
        },
        "it-IT-chan": {
            "id": "UCuPxlWUPz2YS79iwqsAiB9Q",
            "languages": ["it-IT"]
        },
        "pl-PL-chan": {
            "id": "UCPpXRzmjy9BghMGnwF_sq7A",
            "languages": ["pl-PL"]
        },
        "pt-BR-chan":{
            "id": "UCYNPjYq5f0sjRSNZAy99qHA",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UC7Oo1q9m4vf7XrlZgOR95WA",
            "languages": ["ru-RU"]
        },
        "zh-TW-chan":{
            "id": "UCgNBSxvqQMhVYBo-XSnmbqg",
            "languages": ["zh-TW"]
        },
    },
    "Blizzard": {
        "de-DE-chan": {
            "id": "UCm87yTwwLLOfnfiySCGRhBg",
            "languages": ["de-DE"]
        },
        "de-DE-PEGI-chan": {
            "id": "UCm87yTwwLLOfnfiySCGRhBg",
            "languages": ["de-DE-PEGI"]
        },
        "en-US-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["en-US"]
        },
        "en-GB-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["en-GB"]
        },
        "en-EU-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["en-EU"]
        },
        "en-ANZ-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["en-ANZ"]
        },
        "es-ES-chan": {
            "id": "UCNYXB7S9gmAeSZtIKTVl82w",
            "languages": ["es-ES"]
        },
        "es-MX-chan": {
            "id": "UCNYXB7S9gmAeSZtIKTVl82w",
            "languages": ["es-MX"]
        },
        "fr-FR-chan": {
            "id": "UCJZNv5xXuppPxpg8cztq8nA",
            "languages": ["fr-FR"]
        },
        "ko-KR-chan":{
            "id": "UCEOoYEH7uLJ0Ux566Hy3XmQ",
            "languages": ["ko-KR"]
        },
        "it-IT-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["it-IT"],
        },
        "pl-PL-chan": {
            "id": "UC3GriadTkHBnfgd2UFETGOA",
            "languages": ["pl-PL"],
        },
        "pt-BR-chan":{
            "id": "UCLpMxSNFIlTcwL3ELUMqjCA",
            "languages": ["pt-BR"]
        },
        "ru-RU-chan":{
            "id": "UCO8LH6ixpMXKww3OZin042g",
            "languages": ["ru-RU"]
        },
    },
}

# Privacy exceptions:

def privacy_exceptions(channel_language, selected_franchise, selected_privacy):
    return_value = selected_privacy
    if channel_language == "en-GB" and not selected_franchise == "Overwatch" and selected_privacy == "Public":
        return_value = "Unlisted"
    if channel_language == "en-EU" and not selected_franchise == "Overwatch" and selected_privacy == "Public":
        return_value = "Unlisted"
    if channel_language == "it-IT" and selected_franchise == "Blizzard" and selected_privacy == "Public":
        return_value = "Unlisted"
    if channel_language == "pl-PL" and selected_franchise == "Blizzard" and selected_privacy == "Public":
        return_value = "Unlisted"
    if channel_language == "en-ANZ" and selected_privacy == "Public":
        return_value = "Unlisted"
    return return_value

    # Privacy warning:

def privacy_approved(privacy_check_popup_window):
    privacy_variable.set(privacy_list[2]) # Public
    privacy_check_popup_window.destroy()

def privacy_selected_public(new_value):
    global privacy_variable
    if privacy_variable.get() == "Public":
        privacy_variable.set(privacy_list[0])
        privacy_popup()

def privacy_popup():
    #This function defines a popup window that opens when you click the red button in GUI
    privacy_check_popup_window = Toplevel()
    privacy_check_popup_window.geometry(get_main_window_of_gui_postion()) #this line makes popup open at the place of main window
    privacy_check_popup_window.wm_attributes("-topmost", 1) #this line makes popup open on top of main window
    privacy_check_popup_window.wm_title("Privacy Check")
    post_comment_input_field = Label(privacy_check_popup_window, text = "Are you sure that this content should go live immediately? \n\n Click 'No' to move it back to Private. \n")
    post_comment_input_field.pack()
    Yes_button = Button(privacy_check_popup_window, text="Yes", width = 10, height = 1, command = lambda: privacy_approved(privacy_check_popup_window))
    Yes_button.pack()
    No_button = Button(privacy_check_popup_window, text="No", width = 10, height = 1, command = lambda: privacy_check_popup_window.destroy())
    No_button.pack()
    privacy_check_popup_window.mainloop()

def get_main_window_of_gui_postion():
    x_coordinate = main_window_of_gui.winfo_x()
    y_coordinate = main_window_of_gui.winfo_y()
    return ("350x150" + "+" + (str(x_coordinate) + "+" + str(y_coordinate)))

def add_comment_button_clicked(place_holder_selected_media, post_comment_input_field, add_post_comment_popup_window):
    the_row_of_the_comment = find_the_row_of_the_next_empty_cell(place_holder_selected_media) - 1
    wb = load_workbook(generate_the_excel_file_name_with_current_month_and_year_in_name()) #open excel file
    wb.active = place_holder_selected_media #define active excel tab depending on wich media i've chosen
    ws = wb.active #define the fact that i'm working with currently active tab
    cell_to_paste_comment_to = ws['c' + str(the_row_of_the_comment)]  #define a variable for a cell where i will copy my data to
    cell_to_paste_comment_to.value = post_comment_input_field.get("1.0","end")
    wb.save(generate_the_excel_file_name_with_current_month_and_year_in_name())
    add_post_comment_popup_window.destroy()

# Video duration:

from moviepy.editor import VideoFileClip

def get_video_duration():
    for videos in all_videos:
        if videos[0] in video.values():
            clip = VideoFileClip((videos[1] + "\\" + videos[0]).replace("\\","/"))
            duration = float(clip.duration)
            clip.reader.close()
            clip.audio.reader.close_proc()
            return(duration)

# Script for Excel file being created:

def excel_spreadsheet_creation(selected_franchise, selected_privacy, selected_usage_policy, selected_enable_content_id, selected_match_policy, selected_notify_subscribers_option):
    default_usage_policy = "Track Worldwide"
    default_category = "Gaming"
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    # First row:
    column = 1  
    ws.cell(row=1, column=column, value="filename")
    column += 1
    ws.cell(row=1, column=column, value="title")
    column += 1
    ws.cell(row=1, column=column, value="channel")
    column += 1
    ws.cell(row=1, column=column, value="category")
    column += 1
    ws.cell(row=1, column=column, value="privacy")
    column += 1
    if selected_usage_policy != default_usage_policy:
        ws.cell(row=1, column=column, value="usage_policy")
        column += 1
    if get_video_duration() > 5:
        ws.cell(row=1, column=column, value="enable_content_id")
        column += 1
        ws.cell(row=1, column=column, value="match_policy")
        column += 1
    ws.cell(row=1, column=column, value="notify_subscribers")
    column += 1
    if check_if_thumbnails_exists():
        ws.cell(row=1, column=column, value="custom_thumbnail")
        column += 1
    elif len(all_images) == 1:
        ws.cell(row=1, column=column, value="custom_thumbnail")
        column += 1
    if check_if_subtitles_exist():
        ws.cell(row=1, column=column, value="caption_file")
        column += 1
        ws.cell(row=1, column=column, value="caption_language") 
    # All other rows:   
    row = 2 
    for channel in franchises[selected_franchise].values():
        channel_language = channel["languages"][0]
        if not video[channel_language]:
            continue
        yield video[channel_language]
        column = 1
        if ignore_deDE_PEGI_variable.get() == "Yes" and "dedepegi" in video[channel_language].lower().replace("-","").replace("_",""):
            continue
        else:
            ws.cell(row=row, column=column, value=video[channel_language])
            column += 1
            ws.cell(row=row, column=column, value=video[channel_language])
            column += 1
            ws.cell(row=row, column=column, value=channel["id"])
            column += 1
            ws.cell(row=row, column=column, value=default_category)
            column += 1
            ws.cell(row=row, column=column, value=privacy_exceptions(channel_language, selected_franchise, selected_privacy))
            column += 1
            if selected_usage_policy != default_usage_policy:
                ws.cell(row=row, column=column, value=selected_usage_policy)
                column += 1
            if get_video_duration() > 5:
                ws.cell(row=row, column=column, value=selected_enable_content_id)
                column += 1
                ws.cell(row=row, column=column, value=selected_match_policy)
                column += 1
            ws.cell(row=row, column=column, value=selected_notify_subscribers_option)
            column += 1
            if check_if_thumbnails_exists():
                ws.cell(row=row, column=column, value=image[channel_language])
                column += 1
            elif len(all_images) == 1:
                ws.cell(row=row, column=column, value=all_images[0][0])
                column += 1 
            if check_if_subtitles_exist():
                ws.cell(row=row, column=column, value=subtitle[channel_language])
                column += 1
                ws.cell(row=row, column=column, value=file_language[channel_language]["caption_language"])
                column += 1
            row = row + 1
    
    videos_in_folder = "Number of video files in the folder: " + str(len(all_videos))
    videos_analysed = "Number of video files covered by the script: " + (str(row - 2))
    insert_log(videos_in_folder)
    insert_log(videos_analysed) 
    # Save: 
    selected_output_path = save_folder()
    wb.save(selected_output_path)

"""S:\Community\Public\Public\For Maciej\Excel Template"""

#User interface goes here:


#below is the code for "static" part of GUI
main_window_of_gui = Tk()
main_window_of_gui.title("YouTube Batch Upload")
main_window_of_gui.geometry("400x500")

#folder selection window

text_in_gui = Label(main_window_of_gui, text="")
text_in_gui.grid(row=0, columnspan=2)

text_in_gui = Label(main_window_of_gui, text = "Select folder:")
text_in_gui.grid(row = 1, column = 0)

def browse_button():
    global all_files
    global all_videos
    global all_subtitles
    global all_images
    global folder_was_selected
    selected_path = filedialog.askdirectory().replace("\\","/")
    selected_path = os.path.normpath(selected_path)
    insert_log('Selected path:')
    if selected_path != ".":
        folder_was_selected = 1
        insert_log(selected_path)
        insert_log("")
    elif folder_was_selected == 0:
        insert_log("No folder selected")
    all_files = filename_extractor(selected_path)
    all_videos = identify_videos(all_files)
    all_subtitles = identify_subtitles(all_files)
    all_images = identify_images(all_files)
    check_image_size(all_images)
    video_dictionary()
    subtitle_and_image_dictionary()
    check_if_videos_subtitles_and_images_match()


button2 = Button(text="Browse...", command=browse_button).grid(row=1, column=1)

text_in_gui = Label(main_window_of_gui, text="")
text_in_gui.grid(row=2, columnspan=2)

#youtube channel selection:

text_in_gui = Label(main_window_of_gui, text = "YouTube channel:")
text_in_gui.grid(row = 3, column = 0)

ip_selector_variable = StringVar()
ip_list = ["Hearthstone", "Heroes", "Overwatch", "Diablo", "StarCraft", "WoW", "Blizzard"]
ip_selector = OptionMenu(main_window_of_gui, ip_selector_variable, *ip_list)
ip_selector_variable.set(ip_list[0])
ip_selector.config(width=17)
ip_selector.grid(row = 3, column = 1)

#privacy selection:

text_in_gui = Label(main_window_of_gui, text = "Privacy:")
text_in_gui.grid(row = 4, column = 0)


privacy_variable = StringVar()
privacy_list = ["Private", "Unlisted", "Public"]
privacy = OptionMenu(main_window_of_gui, privacy_variable, *privacy_list, command = privacy_selected_public)
privacy_variable.set(privacy_list[0])
privacy.config(width=17)
privacy.grid(row = 4, column = 1)


#usage policy selection:

text_in_gui = Label(main_window_of_gui, text = "Usage policy:")
text_in_gui.grid(row = 5, column = 0)

usage_policy_variable = StringVar()
usage_policy_list = ["Track Worldwide", "BlizzCon VT Policy", "Standard Track", "Block Worldwide", "BlizzCon 2018"]
usage_policy = OptionMenu(main_window_of_gui, usage_policy_variable, *usage_policy_list)
usage_policy_variable.set(usage_policy_list[0])
usage_policy.config(width=17)
usage_policy.grid(row = 5, column = 1)

#enable content ID:

text_in_gui = Label(main_window_of_gui, text = "Enable content ID:")
text_in_gui.grid(row = 6, column = 0)

enable_content_id_variable = StringVar()
enable_content_id_list = ["Yes", "No"]
enable_content_id = OptionMenu(main_window_of_gui, enable_content_id_variable, *enable_content_id_list)
enable_content_id_variable.set(enable_content_id_list[0])
enable_content_id.config(width=17)
enable_content_id.grid(row = 6, column = 1)

#match policy:

text_in_gui = Label(main_window_of_gui, text = "Match policy:")
text_in_gui.grid(row = 7, column = 0)

match_policy_id_variable = StringVar()
match_policy_id_list = ["Track Worldwide", "BlizzCon VT Policy", "Standard Track", "Block Worldwide", "BlizzCon 2018"]
match_policy_id = OptionMenu(main_window_of_gui, match_policy_id_variable, *match_policy_id_list)
match_policy_id_variable.set(match_policy_id_list[0])
match_policy_id.config(width=17)
match_policy_id.grid(row = 7, column = 1)

#notify subscribers:

text_in_gui = Label(main_window_of_gui, text = "Notify subscribers:")
text_in_gui.grid(row = 8, column = 0)

notify_subscribers_id_variable = StringVar()
notify_subscribers_id_list = ["Yes", "No"]
notify_subscribers_id = OptionMenu(main_window_of_gui, notify_subscribers_id_variable, *notify_subscribers_id_list)
notify_subscribers_id_variable.set(notify_subscribers_id_list[0])
notify_subscribers_id.config(width=17)
notify_subscribers_id.grid(row = 8, column = 1)

#ignore deDE-PEGI files:

text_in_gui = Label(main_window_of_gui, text = "Ignore deDE-PEGI files:")
text_in_gui.grid(row = 9, column = 0)

ignore_deDE_PEGI_variable = StringVar()
ignore_deDE_PEGI_list = ["Yes", "No"]
ignore_deDE_PEGI = OptionMenu(main_window_of_gui, ignore_deDE_PEGI_variable, *ignore_deDE_PEGI_list)
ignore_deDE_PEGI_variable.set(ignore_deDE_PEGI_list[0])
ignore_deDE_PEGI.config(width=17)
ignore_deDE_PEGI.grid(row = 9, column = 1)

#text box
text_box = Listbox(main_window_of_gui, height=5)
text_box.grid(column=0, row=12, columnspan=2, sticky=(N,W,E,S))  # columnspan − How many columns widgetoccupies; default 1.
main_window_of_gui.grid_columnconfigure(0, weight=1)
main_window_of_gui.grid_rowconfigure(12, weight=1)

#scroll bar
my_scrollbar = ttk.Scrollbar(main_window_of_gui, orient=VERTICAL, command=text_box.yview)
my_scrollbar.grid(column=2, row=12, sticky=(N,S))
#attaching scroll bar to text box
text_box['yscrollcommand'] = my_scrollbar.set

#inserting the text
def insert_log(text):
    text_box.insert('end', text)
    text_box.see("end")
    

progress=Progressbar(main_window_of_gui,orient=HORIZONTAL, mode='determinate')
progress.grid(row=13, column=0, columnspan=3, sticky=(W,E))

def bar():
    global folder_was_selected
    if folder_was_selected == 1:
        progress.start(interval=None)

        import time
        maximum_videos = len(filter_video(video.values()))
        for i, value in enumerate(generate_excel()):
            if ignore_deDE_PEGI_variable.get() == "Yes" and "dedepegi" in value.lower().replace("-","").replace("_",""):
                insert_log("!Skipped! " + str(value))
            else:
                insert_log("Working on "+ str(value))
                update_progressbar(progress, i+1, maximum_videos)
        final_confirmation()
    else:
        insert_log("No folder selected")
        
def update_progressbar(progressbar, current_value, total_values):
    progress['value']= current_value / total_values * 100
    main_window_of_gui.update_idletasks()

def final_confirmation():
    progress.stop()
    progress['value']= 100

save_folder_was_selected = 0

def save_folder():    
    save_file = filedialog.asksaveasfilename(initialdir = "/",title = "Save File",filetypes = (("Excel files","*.xlsx"),("All files","*.*")))
    global save_folder_was_selected
    if save_file != ".":
        save_folder_was_selected = 1
        insert_log("")
        insert_log("File saved: " + save_file + ".xlsx")
        insert_log("")
        print(save_file)
    elif save_folder_was_selected == 0:
        insert_log("No folder selected")
        print(save_file + "error")
    if save_folder_was_selected == 1:
        return(save_file + ".xlsx")

#generate button:

def generate_excel():
    return excel_spreadsheet_creation(
        ip_selector_variable.get(),
        privacy_variable.get(),
        usage_policy_variable.get(),
        enable_content_id_variable.get(),
        match_policy_id_variable.get(),
        notify_subscribers_id_variable.get())

def generate_click():
    bar()

text_in_gui = Label(main_window_of_gui, text="")
text_in_gui.grid(row=10, columnspan=2)

button3 = Button(text="Generate the file!", command=generate_click).grid(row=11, columnspan=2)


mainloop()

#start the main window of GUI
main_window_of_gui.mainloop()



#def root_folder(filenames):
 #   result = ""
  #  for any_file, root in filenames:
